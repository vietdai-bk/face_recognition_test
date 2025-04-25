import cv2
import face_recognition
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
from PIL import Image, ImageTk
import logging

EXCEL_FILE = "attendance.xlsx"

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Name", "Timestamp"])
        wb.save(EXCEL_FILE)
    return load_workbook(EXCEL_FILE)

def recognize_faces(known_face_encodings, known_face_names, video_label, app, update_status, stop_event, attendance_status):
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        app.root.after(0, lambda: update_status("Không thể truy cập camera!"))
        logger.error("Không thể mở camera")
        return

    wb = init_excel()
    ws = wb.active
    face_confirmations = {name: 0 for name in known_face_names}
    recognized_names = set()

    update_status("Bắt đầu nhận diện khuôn mặt.")
    logger.debug("Bắt đầu vòng lặp nhận diện")

    while not stop_event.is_set():
        ret, frame = cap.read()
        if not ret:
            app.root.after(0, lambda: update_status("Không thể đọc frame từ camera!"))
            logger.error("Không thể đọc frame từ camera")
            break

        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame, model="hog")
        face_encodings = face_recognition.face_encodings(rgb_frame, known_face_locations=face_locations)

        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=0.4)
            name = "Unknown"

            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            if len(face_distances) > 0:
                best_match_index = np.argmin(face_distances)
                if matches[best_match_index]:
                    name = known_face_names[best_match_index]
                    face_confirmations[name] += 1

                    if face_confirmations[name] >= 15 and name not in recognized_names:
                        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        ws.append([name, timestamp])
                        wb.save(EXCEL_FILE)
                        recognized_names.add(name)
                        attendance_status[name] = True
                        app.root.after(0, lambda n=name: update_status(f"Đã ghi {n} vào attendance.xlsx."))
                        app.root.after(0, lambda: app.update_attendance_status())
                        logger.debug(f"Đã ghi {name} vào Excel")
                else:
                    logger.debug("Gương mặt không khớp với bất kỳ dữ liệu nào đã biết")

            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

        cv2.imshow("Nhận diện khuôn mặt", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            stop_event.set()
            break

    try:
        cap.release()
        cv2.destroyAllWindows()
        app.root.after(0, lambda: update_status("Đã dừng nhận diện."))
        logger.debug("Đã dừng nhận diện và giải phóng camera")
    except Exception as e:
        logger.error(f"Lỗi khi giải phóng camera: {str(e)}")
